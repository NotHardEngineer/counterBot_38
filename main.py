import telebot
from telebot import types
from telebot import apihelper
import pandas as pd
from urllib.request import urlopen
from openpyxl import load_workbook
import configparser

config_obj = configparser.ConfigParser()
config_obj.read("сonfigfile.ini", encoding='utf-8-sig')
token = config_obj["program_settings"]["token"]
tablepath = config_obj["program_settings"]["tablepath"]
coldWaterSheetName = config_obj["table_info"]["coldwatersheetname"]
hotWaterSheetName = config_obj["table_info"]["hotwatersheetname"]
electricitySheetName = config_obj["table_info"]["electricitysheetname"]
gasSheetName = config_obj["table_info"]["gassheetname"]

gasLength = int(config_obj["counter_info"]["gaslength"])
coldWaterLength = int(config_obj["counter_info"]["coldwaterlength"])
hotWaterLength = int(config_obj["counter_info"]["hotwaterlength"])
electricityLenght = int(config_obj["counter_info"]["electricitylenght"])

bot = telebot.TeleBot(token)
apihelper.proxy = {'http': 'http://88.204.154.155:8080'}

mouths = ['январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь',
          'октябрь', 'ноябрь', 'декабрь']


@bot.message_handler(commands=['start'])
def start_message(message):
    bot.send_message(message.chat.id,'Привет')
    date = urlopen('http://just-the-time.appspot.com/').read().strip().decode('utf-8')
    curMouthNumber = date[5:7]
    curMouth = mouths[int(curMouthNumber) - 1]


    if not(curMouth in tablepath.lower()):
        bot.send_message(message.chat.id, 'ВНИМАНИЕ: СЕЙЧАС ' + curMouth +
                         " ВОЗМОЖНО НУЖНО ПОМЕНЯТЬ ТАБЛИЦУ, ИНАЧЕ СТАРЫЕ ЗНАЧЕНИЯ МОГУТ БЫТЬ УТЕРЯНЫ")
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    item1 = types.KeyboardButton("Газ")
    item2 = types.KeyboardButton("ГВС")
    item3 = types.KeyboardButton("ХВС")
    item4 = types.KeyboardButton("Электричество")
    markup.add(item1, item2, item3, item4)
    bot.send_message(message.chat.id, 'Выберите тип счетчика', reply_markup=markup)


@bot.message_handler(content_types=['text'])
def message_reply(message):
    if message.text=="Газ":
        a = telebot.types.ReplyKeyboardRemove()
        mgs = bot.send_message(message.chat.id,("Введите через пробел номер счетчика и показания (не больше " + str(gasLength) + " цифр)"),
                               reply_markup=a)
        bot.register_next_step_handler(mgs, gasCounter)
    elif message.text=="ГВС":
        a = telebot.types.ReplyKeyboardRemove()
        mgs = bot.send_message(message.chat.id, ("Введите показания (не больше " + str(hotWaterLength) + " цифр)"),
                               reply_markup=a)
        bot.register_next_step_handler(mgs, hotWaterCounter)
    elif message.text=="ХВС":
        a = telebot.types.ReplyKeyboardRemove()
        mgs = bot.send_message(message.chat.id, ("Введите показания (не больше " + str(coldWaterLength) + " цифр)"),
                               reply_markup=a)
        bot.register_next_step_handler(mgs, coldWaterCounter)
    elif message.text == "Электричество":
        a = telebot.types.ReplyKeyboardRemove()
        mgs = bot.send_message(message.chat.id, ("Введите показания (не больше " + str(electricityLenght) + " цифр)"),
                               reply_markup=a)
        bot.register_next_step_handler(mgs, electricityCounter)


def gasCounter(message):
    # Чтение сообщения
    chat_id = message.chat.id
    text = message.text
    try:
        nOfCounter, inputData = text.split()
    except ValueError:
        msg = bot.send_message(chat_id, 'Вы должны ввести номер счетчика и его показания через пробел')
        bot.register_next_step_handler(msg, gasCounter)
        return
    if len(inputData) > gasLength:
        msg = bot.send_message(chat_id, 'Вы должны ввести не больше ' + str(gasLength) + " цифр в качестве показаний")
        bot.register_next_step_handler(msg, gasCounter)
        return
    try:
        nOfCounter = int(nOfCounter)
        inputData = float(inputData.replace(",", "."))
    except:
        msg = bot.send_message(chat_id, 'Вы должны ввести число)')
        bot.register_next_step_handler(msg, hotWaterCounter)
        return
    curDay = urlopen('http://just-the-time.appspot.com/').read().strip().decode('utf-8')[8:10]
    try:
        xl = pd.ExcelFile(tablepath)
    except FileNotFoundError:
        bot.send_message(chat_id, "Не найден файл таблицы, данные не записаны")
        return
    sheetList = xl.sheet_names
    # Поиск нужной таблицы
    sheetName = 'placeholder'
    for sheetsName in sheetList:
        if gasSheetName.lower() in sheetsName.lower():
            sheetName = sheetsName
            break
    try:
        sheet = xl.parse(sheetName)
    except NameError:
        bot.send_message(chat_id, "Не найдена таблица для этого счетчика, данные не записаны")
        xl.close()
        return

    # Ищем строку с заголовками и датами
    searchFrame = sheet.loc[sheet[sheet.columns[0]] == '№']
    headerRowIndex = searchFrame.index[0].tolist() + 2
    headerRowList = sheet.fillna(axis=0, method='ffill', limit=3).loc[sheet.index[headerRowIndex]].tolist()
    print(headerRowList)
    for headerIndex in range(len(headerRowList)):
        headerRowList[headerIndex] = str(headerRowList[headerIndex]).lower()
    for i in range(headerRowList.index("1"), headerRowList.index("итого")):
        headerRowList[i] = int(float(headerRowList[i]))

    # Ищем строку нужного счетчика
    searchFrame = sheet.loc[sheet[sheet.columns[0]] == nOfCounter]
    if searchFrame.empty:
        bot.send_message(chat_id, "Не найден счетчик под нужным номером, данные не записаны")
        xl.close()
        return

    writeRow = searchFrame.index[0].tolist() + 2
    print(headerRowList)
    writeCol = headerRowList.index(int(curDay)) + 1

    try:
        wb = load_workbook(tablepath)
        ws = wb[sheetName]
        ws.cell(row=writeRow, column=writeCol).value = inputData
        wb.save(tablepath)
        wb.close()
    except PermissionError:
        bot.send_message(chat_id, "Ошибка доступа к таблице, данные не записаны")
        xl.close()
        return
    xl.close()
    msg = bot.send_message(chat_id, "Данные в таблице!")
    start_message(msg)
    return

def hotWaterCounter(message):
    # Чтение сообщения
    curDay = urlopen('http://just-the-time.appspot.com/').read().strip().decode('utf-8')[8:10]
    chat_id = message.chat.id
    text = message.text
    try:
        nOfCounter, inputData = text.split()
    except ValueError:
        msg = bot.send_message(chat_id, 'Вы должны ввести номер счетчика и его показания')
        bot.register_next_step_handler(msg, hotWaterCounter)
        return
    if len(inputData) > hotWaterLength:
        msg = bot.send_message(chat_id,
                               'Вы должны ввести не больше ' + str(hotWaterLength) + " цифр в качестве показаний")
        bot.register_next_step_handler(msg, hotWaterCounter)
        return
    try:
        nOfCounter = int(nOfCounter)
        inputData = float(inputData.replace(",", "."))
    except:
        msg = bot.send_message(chat_id, 'Вы должны ввести число)')
        bot.register_next_step_handler(msg, hotWaterCounter)
        return
    try:
        xl = pd.ExcelFile(tablepath)
    except FileNotFoundError:
        bot.send_message(chat_id, "Не найден файл таблицы, данные не записаны")
        start_message()
        return
    sheetList = xl.sheet_names
    # Поиск нужной таблицы
    for sheetsName in sheetList:
        if hotWaterSheetName.lower() in sheetsName.lower():
            sheetName = sheetsName
            break
    sheetName = 'placeholder'
    try:
        sheet = xl.parse(sheetName)
    except NameError:
        bot.send_message(chat_id, "Не найдена таблица для этого счетчика")
        xl.close()
        return

    # Ищем строку с заголовками и датами
    searchFrame = sheet.loc[sheet[sheet.columns[0]] == '№']
    headerRowIndex = searchFrame.index[0].tolist() + 2
    headerRowList = sheet.fillna(axis=0, method='ffill', limit=3).loc[sheet.index[headerRowIndex]].tolist()
    for headerIndex in range(len(headerRowList)):
        headerRowList[headerIndex] = str(headerRowList[headerIndex]).lower()
    for i in range(headerRowList.index("итого") + 2, len(headerRowList)):
        try:
            headerRowList[i] = int(headerRowList[i][8:10])
        except ValueError:
            break


    # Ищем строку нужного счетчика
    searchFrame = sheet.loc[sheet[sheet.columns[0]] == nOfCounter]
    if searchFrame.empty:
        bot.send_message(chat_id, "Не найден счетчик под нужным номером, данные не записаны")
        xl.close()
        return

    writeRow = searchFrame.index[0].tolist() + 2
    writeCol = headerRowList.index(int(curDay)) + 1

    try:
        wb = load_workbook(tablepath)
        ws = wb[sheetName]
        ws.cell(row=writeRow, column=writeCol).value = inputData
        wb.save(tablepath)
        wb.close()
    except PermissionError:
        bot.send_message(chat_id, "Ошибка доступа к таблице, данные не записанны")
        xl.close()
        return
    xl.close()
    msg = bot.send_message(chat_id, "Данные в таблице!")
    start_message(msg)
    return


def coldWaterCounter(message):
    # Чтение сообщения
    curDay = urlopen('http://just-the-time.appspot.com/').read().strip().decode('utf-8')[8:10]
    chat_id = message.chat.id
    text = message.text
    try:
        nOfCounter, inputData = text.split()
    except ValueError:
        msg = bot.send_message(chat_id, 'Вы должны ввести номер счетчика и его показания через пробел')
        bot.register_next_step_handler(msg, coldWaterCounter)
        return
    if len(inputData) > coldWaterLength:
        msg = bot.send_message(chat_id,
                               'Вы должны ввести не больше ' + str(coldWaterLength) + " цифр в качестве показаний")
        bot.register_next_step_handler(msg, coldWaterCounter)
        return
    try:
        nOfCounter = int(nOfCounter)
        inputData = float(inputData.replace(",", "."))
    except:
        msg = bot.send_message(chat_id, 'Вы должны ввести число)')
        bot.register_next_step_handler(msg, hotWaterCounter)
        return
    try:
        xl = pd.ExcelFile(tablepath)
    except FileNotFoundError:
        bot.send_message(chat_id, "Не найден файл таблицы, данные не записаны")
        return
    sheetList = xl.sheet_names
    # Поиск нужной таблицы
    sheetName = 'placeholder'
    for sheetsName in sheetList:
        if coldWaterSheetName.lower() in sheetsName.lower():
            sheetName = sheetsName
            break

    try:
        sheet = xl.parse(sheetName)
    except NameError:
        bot.send_message(chat_id, "Не найдена таблица для этого счетчика, данные не записаны")
        xl.close()
        return

    # Ищем строку с заголовками и датами
    searchFrame = sheet.loc[sheet[sheet.columns[0]] == '№']
    headerRowIndex = searchFrame.index[0].tolist() + 2
    headerRowList = sheet.fillna(axis=0, method='ffill', limit=3).loc[sheet.index[headerRowIndex]].tolist()
    for headerIndex in range(len(headerRowList)):
        headerRowList[headerIndex] = str(headerRowList[headerIndex]).lower()
    print(headerRowList)
    for i in range(headerRowList.index("итого") + 2, len(headerRowList) - 1):
        try:
            headerRowList[i] = int(headerRowList[i][8:10])
        except ValueError:
            break

    # Ищем строку нужного счетчика
    searchFrame = sheet.loc[sheet[sheet.columns[0]] == nOfCounter]
    if searchFrame.empty:
        bot.send_message(chat_id, "Не найден счетчик под нужным номером, данные не записаны")
        xl.close()
        return

    writeRow = searchFrame.index[0].tolist() + 2
    writeCol = headerRowList.index(int(curDay)) + 1

    try:
        wb = load_workbook(tablepath)
        ws = wb[sheetName]
        ws.cell(row=writeRow, column=writeCol).value = inputData
        wb.save(tablepath)
        wb.close()
    except PermissionError:
        bot.send_message(chat_id, "Ошибка доступа к таблице, данные не записаны")
        xl.close()
        return
    xl.close()
    msg = bot.send_message(chat_id, "Данные в таблице!")
    start_message(msg)
    return


def electricityCounter(message):
    # Чтение сообщения
    curDay = urlopen('http://just-the-time.appspot.com/').read().strip().decode('utf-8')[8:10]
    chat_id = message.chat.id
    text = message.text
    try:
        nOfCounter, inputData = text.split()
    except ValueError:
        msg = bot.send_message(chat_id, 'Вы должны ввести номер счетчика и его показания через пробел')
        bot.register_next_step_handler(msg, electricityCounter)
        return
    if len(inputData) > electricityLenght:
        msg = bot.send_message(chat_id,
                               'Вы должны ввести не больше ' + str(electricityLenght) + " цифр в качестве показаний")
        bot.register_next_step_handler(msg, electricityCounter)
        return
    try:
        nOfCounter = int(nOfCounter)
        inputData = float(inputData.replace(",", "."))
    except:
        msg = bot.send_message(chat_id, 'Вы должны ввести число)')
        bot.register_next_step_handler(msg, hotWaterCounter)
        return

    curDay = urlopen('http://just-the-time.appspot.com/').read().strip().decode('utf-8')[8:10]
    try:
        xl = pd.ExcelFile(tablepath)
    except FileNotFoundError:
        bot.send_message(chat_id, "Не найден файл таблицы, данные не записаны")
        return
    sheetList = xl.sheet_names
    # Поиск нужной таблицы
    for sheetsName in sheetList:
        if electricitySheetName.lower() in sheetsName.lower():
            sheetName = sheetsName
            break
    sheetName = 'placeholder'
    try:
        sheet = xl.parse(sheetName)
    except NameError:
        bot.send_message(chat_id, "Не найдена таблица для этого счетчика, данные не записаны")
        xl.close()
        return

    # Ищем строку с заголовками и датами
    searchFrame = sheet.loc[sheet[sheet.columns[0]] == '№ счетчика']
    headerRowIndex = searchFrame.index[0].tolist() + 3
    headerRowList = sheet.fillna(axis=0, method='ffill', limit=3).loc[sheet.index[headerRowIndex]].tolist()
    for headerIndex in range(len(headerRowList)):
        headerRowList[headerIndex] = str(headerRowList[headerIndex]).lower()
    for i in range(headerRowList.index("к(тр)") + 1, len(headerRowList) - 1):
        try:
            headerRowList[i] = int(headerRowList[i][8:10])
        except ValueError:
            break

    # Ищем строку нужного счетчика
    searchFrame = sheet.loc[sheet[sheet.columns[0]] == nOfCounter]
    if searchFrame.empty:
        bot.send_message(chat_id, "Не найден счетчик под нужным номером, данные не записаны")
        xl.close()
        return

    writeRow = searchFrame.index[0].tolist() + 2

    writeCol = headerRowList.index(int(curDay)) + 1

    try:
        wb = load_workbook(tablepath)
        ws = wb[sheetName]
        ws.cell(row=writeRow, column=writeCol).value = inputData
        wb.save(tablepath)
        wb.close()
    except PermissionError:
        bot.send_message(chat_id, "Ошибка доступа к таблице, данные не записаны")
        xl.close()
        return
    xl.close()
    msg = bot.send_message(chat_id, "Данные в таблице!")
    start_message(msg)
    return


bot.infinity_polling()